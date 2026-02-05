import tkinter as tk
from tkinter import ttk, messagebox
import requests
import datetime
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
import webbrowser
import warnings

try:
    from openpyxl import Workbook  # type: ignore
    from openpyxl.styles import Font, Alignment, PatternFill  # type: ignore
except (ImportError, ModuleNotFoundError):
    Workbook = None
    Font = Alignment = PatternFill = None

warnings.filterwarnings("ignore", category=UserWarning, module="matplotlib")

API_KEY = "UxDoRKNrs5en8rQeHz9LlcJJ_usQ9Q47"

symbols = [
    "XAU/USD", "ETH/USD", "EUR/USD", "GBP/USD", "USD/JPY",
    "AUD/USD", "NZD/USD", "USD/CAD", "USD/CHF", "EUR/GBP",
    "EUR/JPY", "GBP/JPY", "AUD/JPY", "NZD/JPY", "CHF/JPY",
    "EUR/AUD", "EUR/NZD", "GBP/AUD", "GBP/NZD", "CAD/JPY",
    "EUR/CHF", "GBP/CHF", "AUD/CHF", "NZD/CHF", "CAD/CHF"
]

last_analysis = {}

# إعداد واجهة المستخدم الرئيسية
root = tk.Tk()
root.title("محلل الفوركس الذكي")
root.geometry("800x600")

# إنشاء المتغيرات بعد إنشاء نافذة Tkinter
use_rsi = tk.BooleanVar(value=True)
use_macd = tk.BooleanVar(value=True)
use_bollinger = tk.BooleanVar(value=True)

def fetch_price_data(symbol):
    try:
        if symbol in ["XAU/USD", "ETH/USD"]:
            function = "CURRENCY_EXCHANGE_RATE"
            from_symbol, to_symbol = symbol.split("/")
            url = f"https://www.alphavantage.co/query?function={function}&from_currency={from_symbol}&to_currency={to_symbol}&apikey={API_KEY}"
            response = requests.get(url)
            data = response.json()
            price = float(data["Realtime Currency Exchange Rate"]["5. Exchange Rate"])
            return [price] * 30
        else:
            function = "FX_INTRADAY"
            interval = "15min"
            from_symbol, to_symbol = symbol.split("/")
            url = f"https://www.alphavantage.co/query?function={function}&from_symbol={from_symbol}&to_symbol={to_symbol}&interval={interval}&apikey={API_KEY}"
            response = requests.get(url)
            data = response.json()
            time_series = data.get("Time Series FX (15min)", {})
            if not time_series:
                return None
            sorted_times = sorted(time_series.keys(), reverse=True)
            prices = [float(time_series[time]['4. close']) for time in sorted_times]
            return prices[::-1]
    except Exception as e:
        print(f"Error fetching data: {e}")
        return None

def compute_rsi(prices, period=14):
    prices = np.array(prices)
    deltas = np.diff(prices)
    gain = np.where(deltas > 0, deltas, 0)
    loss = np.where(deltas < 0, -deltas, 0)

    avg_gain = np.convolve(gain, np.ones((period,)) / period, mode='valid')
    avg_loss = np.convolve(loss, np.ones((period,)) / period, mode='valid')

    rs = avg_gain / (avg_loss + 1e-6)
    rsi = 100 - (100 / (1 + rs))
    return np.concatenate(([50] * (len(prices) - len(rsi)), rsi))

def compute_macd(prices):
    prices = np.array(prices)
    ema12 = np.convolve(prices, np.ones(12) / 12, mode='valid')
    ema26 = np.convolve(prices, np.ones(26) / 26, mode='valid')
    min_len = min(len(ema12), len(ema26))
    macd_line = ema12[-min_len:] - ema26[-min_len:]
    signal_line = np.convolve(macd_line, np.ones(9) / 9, mode='valid')
    return macd_line[-len(signal_line):], signal_line

def compute_bollinger_bands(prices, period=20):
    prices = np.array(prices)
    sma = np.convolve(prices, np.ones(period) / period, mode='valid')
    std = np.array([np.std(prices[i - period:i]) for i in range(period, len(prices) + 1)])
    upper = sma + (2 * std)
    lower = sma - (2 * std)
    return sma, upper, lower

def calculate_signals(prices):
    signals = []
    rsi_val = macd_val = None

    if use_rsi.get():
        rsi = compute_rsi(prices)
        rsi_val = rsi[-1]
        if rsi_val < 30:
            signals.append("شراء")
        elif rsi_val > 70:
            signals.append("بيع")

    if use_macd.get():
        macd, signal = compute_macd(prices)
        if len(macd) > 0 and len(signal) > 0:
            macd_val = macd[-1]
            if macd[-1] > signal[-1]:
                signals.append("شراء")
            elif macd[-1] < signal[-1]:
                signals.append("بيع")

    if use_bollinger.get():
        sma, upper, lower = compute_bollinger_bands(prices)
        if prices[-1] < lower[-1]:
            signals.append("شراء")
        elif prices[-1] > upper[-1]:
            signals.append("بيع")

    if "شراء" in signals and "بيع" not in signals:
        return "شراء", rsi_val, macd_val
    elif "بيع" in signals and "شراء" not in signals:
        return "بيع", rsi_val, macd_val
    else:
        return "محايد", rsi_val or 0, macd_val or 0

def plot_data(prices):
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.plot(prices, label='السعر')
    if use_rsi.get():
        rsi = compute_rsi(prices)
        trimmed_prices = prices[-len(rsi):]
        ax.plot(trimmed_prices, label='السعر (معدل)')
        ax2 = ax.twinx()
        ax2.plot(rsi, color='orange', alpha=0.3, label='RSI')
        ax2.scatter(range(len(rsi)), rsi, color='orange', marker='x', facecolors='none', label='RSI نقاط')
    ax.set_title("تحليل السعر والمؤشرات الفنية")
    ax.set_xlabel("الزمن")
    ax.set_ylabel("السعر")
    fig.legend()
    return fig

def export_to_excel(analysis_results):
    if Workbook is None:
        messagebox.showerror("خطأ", "مكتبة openpyxl غير مثبتة. يرجى تثبيتها باستخدام: pip install openpyxl")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "تحليل الفوركس"

    headers = ["الرمز", "التوصية", "RSI", "MACD"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for symbol, data in analysis_results.items():
        ws.append([symbol, data["signal"], data["rsi"], data["macd"]])

    file_name = f"Forex_Analysis_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(file_name)
    messagebox.showinfo("تم التصدير", f"تم تصدير التحليل إلى الملف: {file_name}")

def analyze_all_symbols():
    last_analysis.clear()
    progress = ttk.Progressbar(root, mode="determinate", maximum=len(symbols))
    progress.pack(pady=5)
    for i, symbol in enumerate(symbols):
        prices = fetch_price_data(symbol)
        if prices:
            signal, rsi_val, macd_val = calculate_signals(prices)
            last_analysis[symbol] = {
                "signal": signal,
                "rsi": round(rsi_val, 2),
                "macd": round(macd_val, 4)
            }
        else:
            last_analysis[symbol] = {
                "signal": "بيانات غير متوفرة",
                "rsi": 0,
                "macd": 0
            }
        progress['value'] = i + 1
        root.update_idletasks()
    messagebox.showinfo("تم التحليل", "تم تحليل جميع الأزواج.")

btn_analyze_all = tk.Button(root, text="تحليل جميع الأزواج", command=analyze_all_symbols)
btn_analyze_all.pack(pady=10)

btn_export = tk.Button(root, text="تصدير إلى Excel", command=lambda: export_to_excel(last_analysis))
btn_export.pack(pady=10)

root.mainloop()
